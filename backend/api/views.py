from api.utils.activity import log_login_activity, AuditLogMixin
from api.utils.request_utils import get_client_ip
from datetime import timedelta
from rest_framework import viewsets, permissions, status, filters
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.decorators import action
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.views import APIView
from django.conf import settings
from django.contrib.auth import get_user_model, password_validation
from django.contrib.auth.tokens import (
    PasswordResetTokenGenerator,
    default_token_generator,
)
from django.http import FileResponse, Http404, HttpResponse
from django.db import transaction
from django.db.models import Count, F, Q
from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from zoneinfo import ZoneInfo
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    PageBreak,
    Spacer,
    KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus.doctemplate import LayoutError
from xml.sax.saxutils import escape
from collections import Counter
from .services import update_overdue_actions_if_needed


from .serializers import (
    MyTokenObtainPairSerializer,
    UserSerializer,
    DocumentosLGPDSerializer,
    ChecklistSerializer,
    InventarioDadosSerializer,
    RiskSerializer,
    ActionPlanSerializer,
    MonitoringActionSerializer,
    IncidentSerializer,
    CalendarEventSerializer,
    LoginActivitySerializer,
    UserActivityLogSerializer,
)
from .models import (
    User,
    DocumentosLGPD,
    Checklist,
    InventarioDados,
    Risk,
    ActionPlan,
    MonitoringAction,
    Incident,
    LikelihoodItem,
    ImpactItem,
    CalendarEvent,
    LoginActivity,
    UserActivityLog,
)
from .permissions import (
    IsRoleAdmin,
    IsRoleDPO,
    IsAuthenticatedReadOnly,
    IsAdminOrDPO,
    SimpleRolePermission,
)
from .pagination import DefaultPagination

from .utils.email import send_html_email

import csv
import datetime

try:
    from rest_framework_simplejwt.tokens import RefreshToken
    from rest_framework_simplejwt.exceptions import TokenError

    SIMPLEJWT_BLACKLIST_AVAILABLE = True
except Exception:
    SIMPLEJWT_BLACKLIST_AVAILABLE = False


# View para o login JWT
class MyTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]
    serializer_class = MyTokenObtainPairSerializer


# Helpers p/ cookie httpOnly do refresh
# ---------------------------------------------------------
REFRESH_COOKIE_NAME = "refresh_token"


def _refresh_cookie_kwargs(request=None):
    """
    Define flags do cookie de refresh com detecção automática de ambiente.

    - Em localhost: permite HTTP e SameSite=Lax.
    - Em ambiente HTTPS (túnel/produção): força Secure + SameSite=None + domínio do túnel.
    """
    lifetime = settings.SIMPLE_JWT.get("REFRESH_TOKEN_LIFETIME", timedelta(minutes=15))

    # Detecta host da requisição se disponível
    host = None
    if request is not None:
        host = request.get_host().split(":")[0]
    elif hasattr(settings, "TUNNEL_URL"):
        from urllib.parse import urlparse

        parsed = urlparse(settings.TUNNEL_URL)
        host = parsed.hostname

    is_local = host in ("127.0.0.1", "localhost")

    opts = dict(
        max_age=int(lifetime.total_seconds()),
        httponly=True,
        path="/api/v1/auth/",
    )

    if is_local:
        opts.update(secure=False, samesite="Lax", domain="localhost")
    else:
        cookie_domain = getattr(settings, "COOKIE_DOMAIN", None)
        opts.update(secure=True, samesite="None")
        if cookie_domain:
            opts["domain"] = cookie_domain

    return opts


def _set_refresh_cookie(resp, refresh_str: str, request=None):
    resp.set_cookie(REFRESH_COOKIE_NAME, refresh_str, **_refresh_cookie_kwargs(request))
    return resp


def _clear_refresh_cookie(resp, request=None):
    opts = _refresh_cookie_kwargs(request)
    resp.delete_cookie(
        REFRESH_COOKIE_NAME,
        path=opts.get("path", "/api/v1/auth/"),
        domain=opts.get("domain") or None,
        samesite=opts.get("samesite", "Lax"),
    )
    return resp


# ---------------------------------------------------------
# 1) Login que SETA cookie httpOnly com refresh
#    (mantém o body padrão com access/refresh para compatibilidade)
# ---------------------------------------------------------
class CookieTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]
    serializer_class = MyTokenObtainPairSerializer  # seu serializer custom

    @method_decorator(never_cache)
    def post(self, request, *args, **kwargs):
        res = super().post(request, *args, **kwargs)

        # → Obtém o usuário autenticado
        email = request.data.get("email")
        user = User.objects.filter(email=email).first()

        # → REGISTRO OFICIAL DO LOGIN (CDU14)
        if user:
            print(">> LOGIN LOG DISPARADO")
            log_login_activity(request, user)

        refresh = res.data.get("refresh")
        if refresh:
            _set_refresh_cookie(res, refresh, request)
        # Se quiser não enviar o refresh no body em prod:
        # if not settings.DEBUG:
        #     res.data.pop("refresh", None)
        return res


# ---------------------------------------------------------
# 2) Refresh que LÊ o cookie httpOnly e devolve novo access
#    - Se ROTATE_REFRESH_TOKENS=True, rotaciona o refresh e atualiza o cookie
# ---------------------------------------------------------
class CookieTokenRefreshView(APIView):
    permission_classes = [AllowAny]

    @method_decorator(never_cache)
    def post(self, request):
        raw = request.COOKIES.get(REFRESH_COOKIE_NAME)
        if not raw:
            return Response(
                {"detail": "Refresh ausente."}, status=status.HTTP_401_UNAUTHORIZED
            )

        try:
            current_rt = RefreshToken(raw)
        except TokenError:
            return Response(
                {"detail": "Refresh inválido."}, status=status.HTTP_401_UNAUTHORIZED
            )

        rotate = settings.SIMPLE_JWT.get("ROTATE_REFRESH_TOKENS", False)
        blacklist_after = (
            settings.SIMPLE_JWT.get("BLACKLIST_AFTER_ROTATION", False)
            and SIMPLEJWT_BLACKLIST_AVAILABLE
        )

        # Gera access
        access_str = str(current_rt.access_token)

        new_refresh_str = raw
        if rotate:
            # Blacklist do anterior (se habilitado)
            if blacklist_after:
                try:
                    current_rt.blacklist()
                except Exception:
                    pass

            # Emite um novo refresh para o mesmo usuário
            user_id = current_rt.get("user_id")
            UserModel = get_user_model()
            try:
                user = UserModel.objects.get(pk=user_id)
            except UserModel.DoesNotExist:
                return Response(
                    {"detail": "Usuário não encontrado."},
                    status=status.HTTP_401_UNAUTHORIZED,
                )

            new_rt = RefreshToken.for_user(user)
            new_refresh_str = str(new_rt)
            access_str = str(new_rt.access_token)

        resp = Response({"access": access_str})
        if rotate:
            _set_refresh_cookie(resp, new_refresh_str, request)
        return resp


# ---------------------------------------------------------
# 3) Logout: apaga cookie e (se possível) coloca refresh na blacklist
# ---------------------------------------------------------
class CookieTokenLogoutView(APIView):
    permission_classes = [AllowAny]

    @method_decorator(never_cache)
    def post(self, request):
        raw = request.COOKIES.get(REFRESH_COOKIE_NAME)
        resp = Response({"detail": "OK"})
        if raw and SIMPLEJWT_BLACKLIST_AVAILABLE:
            try:
                RefreshToken(raw).blacklist()
            except TokenError:
                pass
        _clear_refresh_cookie(resp, request)
        return resp


# ViewSet para o modelo User
class UserViewSet(AuditLogMixin, viewsets.ModelViewSet):
    serializer_class = UserSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    queryset = User.objects.none()
    audit_module = "users"

    def _log_activity(self, request, operation, obj=None, result="SUCCESS", details=""):
        """Compatibiliza com sua model UserActivityLog atual."""
        user = request.user if request.user.is_authenticated else None
        email = getattr(user, "email", None)

        UserActivityLog.objects.create(
            usuario=user,
            email=email,
            modulo=self.audit_module,
            operacao=operation.lower(),  # create/update/delete/read
            registro_id=str(getattr(obj, "pk", "")) if obj else None,
            detalhe=details[:5000],
            resultado=result.lower(),
            ip=get_client_ip(request),
        )

    @action(
        detail=False,
        methods=["get"],
        permission_classes=[permissions.IsAuthenticated],
        url_path="dpo",
    )
    def dpo(self, request):
        user = User.objects.filter(role__iexact="dpo").order_by("id").first()
        if not user:
            return Response(
                {"detail": "DPO não encontrado."}, status=status.HTTP_404_NOT_FOUND
            )
        data = self.get_serializer(user, context={"request": request}).data
        return Response(data)

    def get_queryset(self):
        user = self.request.user
        if not (user and user.is_authenticated):
            return User.objects.none()

        # admin vê todos
        if user.is_superuser or getattr(user, "role", "") == "admin":
            show_inactive = self.request.query_params.get("show_inactive")
            # Mostrar usuários desativados somente se show_inactive=1
            if show_inactive == "1":
                qs = User.objects.filter(is_active=False).order_by("email")
            else:
                qs = User.objects.filter(is_active=True).order_by("email")
            q = self.request.query_params.get("q")  # ex.: ?q=joao
            if q:
                qs = qs.filter(
                    Q(email__icontains=q)
                    | Q(first_name__icontains=q)
                    | Q(last_name__icontains=q)
                )
            return qs

        # não-admin:
        if self.action == "list":
            return User.objects.none()  # não expõe diretório de usuários
        if self.action == "retrieve":
            return User.objects.filter(pk=user.pk)  # só o próprio
        return User.objects.none()

    def get_permissions(self):
        # apenas admin pode criar/editar/apagar/listar
        if self.action in ["create", "update", "partial_update", "destroy", "list"]:
            return [IsRoleAdmin()]
        # retrieve e actions custom (me): autenticado
        return [permissions.IsAuthenticated()]

    # Sobrescreve o método perform_create para definir a senha do usuário
    def perform_create(self, serializer):
        serializer.save()

    # Sobrescreve o método perform_update para definir a senha do usuário
    def perform_update(self, serializer):
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        try:
            user = self.get_object()

            # 1. Tentativa inválida – admin tentando desativar a si mesmo
            if user.pk == request.user.pk:
                UserActivityLog.objects.create(
                    usuario=request.user,
                    modulo="users",
                    operacao="deactivate",
                    registro_id=user.pk,
                    detalhe="Tentativa inválida: usuário tentou desativar a si mesmo.",
                    resultado="invalid",
                    ip=request.META.get("REMOTE_ADDR"),
                )

                return Response(
                    {"detail": "Você não pode desativar a si mesmo."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # 2. DESATIVAÇÃO NORMAL (soft delete)
            user.is_active = False
            user.save(update_fields=["is_active"])

            # Registro de sucesso
            UserActivityLog.objects.create(
                usuario=request.user,
                modulo="users",
                operacao="deactivate",
                registro_id=user.pk,
                detalhe=f"Usuário {user.email} desativado.",
                resultado="success",
                ip=request.META.get("REMOTE_ADDR"),
            )

            return Response(
                {"detail": "Usuário desativado com sucesso."},
                status=status.HTTP_200_OK,
            )

        except Exception as exc:
            # 3. Registro de ERRO interno (qualquer outra falha)
            UserActivityLog.objects.create(
                usuario=request.user,
                modulo="users",
                operacao="deactivate",
                registro_id=None,
                detalhe=f"Erro ao tentar desativar usuário: {exc}",
                resultado="error",
                ip=request.META.get("REMOTE_ADDR"),
            )

            # re-levanta o erro para o DRF tratar
            raise

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[IsRoleAdmin],
    )
    def reactivate(self, request, pk=None):
        try:
            user = get_object_or_404(User, pk=pk)

            # Impedir que um usuário reative a si mesmo? (opcional)
            # Mas geralmente permitido, então não aplicamos bloqueio.

            user.is_active = True
            user.save(update_fields=["is_active"])

            # Registrar sucesso
            UserActivityLog.objects.create(
                usuario=request.user,
                modulo="users",
                operacao="reactivate",
                registro_id=user.pk,
                detalhe=f"Usuário {user.email} reativado.",
                resultado="success",
                ip=request.META.get("REMOTE_ADDR"),
            )

            return Response({"detail": "Usuário reativado com sucesso."})

        except Exception as exc:
            # Registrar erro inesperado
            UserActivityLog.objects.create(
                usuario=request.user,
                modulo="users",
                operacao="reactivate",
                registro_id=None,
                detalhe=f"Erro ao tentar reativar usuário: {exc}",
                resultado="error",
                ip=request.META.get("REMOTE_ADDR"),
            )
            raise

    # GET /users/me/ -> dados do próprio usuário
    @action(
        detail=False,
        methods=["get", "patch"],
        permission_classes=[permissions.IsAuthenticated],
        url_path="me",
    )
    def me(self, request):
        user = request.user

        if request.method.lower() == "get":
            serializer = self.get_serializer(user, context={"request": request})
            return Response(serializer.data)

        # Delega tudo ao serializer (avatar, remove_avatar, current_password/password, etc.)
        serializer = self.get_serializer(
            instance=user,
            data=request.data,
            partial=True,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)

        # Detecta se haverá troca de senha antes de salvar
        password_changed = bool(serializer.validated_data.get("password"))

        with transaction.atomic():
            serializer.save()

        # Blacklist opcional do refresh token ao trocar senha
        if password_changed and SIMPLEJWT_BLACKLIST_AVAILABLE:
            raw_refresh = request.data.get("refresh")
            if raw_refresh:
                try:
                    RefreshToken(raw_refresh).blacklist()
                except TokenError:
                    # token inválido/expirado/ausente -> ignore
                    pass

        data_out = self.get_serializer(user, context={"request": request}).data
        data_out["reauth_required"] = bool(password_changed)
        return Response(data_out, status=status.HTTP_200_OK)

    @action(
        detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated]
    )
    def resend_welcome(self, request, pk=None):
        """
        Reenvia o e-mail de boas-vindas para o usuário alvo.
        Restrito a admin/superuser.
        """
        if not (
            request.user.is_superuser or getattr(request.user, "role", "") == "admin"
        ):
            return Response(
                {"detail": "Sem permissão."}, status=status.HTTP_403_FORBIDDEN
            )

        user = self.get_object()
        if not user.email:
            return Response(
                {"detail": "Usuário sem e-mail."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)

            frontend = (
                getattr(settings, "FRONTEND_URL", None)
                or getattr(settings, "FRONTEND_BASE_URL", None)
                or "http://127.0.0.1:5173"
            ).rstrip("/")

            reset_url = f"{frontend}/definir-senha?uid={uid}&token={token}"

            expires_seconds = int(
                getattr(settings, "PASSWORD_RESET_TIMEOUT", 60 * 60 * 24 * 3)
            )
            expires_hours = expires_seconds // 3600

            ctx = {
                "first_name": user.first_name or "",
                "reset_url": reset_url,
                "expires_hours": expires_hours,
            }

            sent = send_html_email(
                subject="Bem-vindo(a) ao Camaleão — defina sua senha",
                to_email=user.email,
                template_name="emails/welcome",
                context=ctx,
            )

            if sent:
                return Response({"detail": "E-mail de boas-vindas reenviado."})
            else:
                return Response(
                    {"detail": "Falha ao enviar o e-mail."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except Exception as exc:
            import logging

            logger = logging.getLogger(__name__)
            logger.exception(f"Erro ao reenviar welcome para {user.email}: {exc}")
            return Response(
                {"detail": "Erro interno ao enviar o e-mail."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DocumentosLGPDViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = DocumentosLGPD.objects.all().order_by("-created_at")
    serializer_class = DocumentosLGPDSerializer
    permission_classes = [SimpleRolePermission]
    audit_module = "documentos"

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["dimensao", "criticidade", "status"]

    OWN_FIELD = "criado_por"
    ROLE_PERMS = {
        # leitura liberada para todos autenticados
        "list": {"admin": "any", "dpo": "any", "user": "any"},
        "retrieve": {"admin": "any", "dpo": "any", "user": "any"},
        # escrita só admin/dpo
        "create": {"admin": "any", "dpo": "any"},
        "update": {"admin": "any", "dpo": "any"},
        "partial_update": {"admin": "any", "dpo": "any"},
        "destroy": {"admin": "any", "dpo": "any"},
        # ações extras
        "upload": {"admin": "any", "dpo": "any"},
        "choices": {"admin": "any", "dpo": "any", "user": "any"},
        # fallback (opcional)
        "*": {"admin": "any", "dpo": "any"},
    }

    @action(detail=False, methods=["get"])
    def choices(self, request):
        m = DocumentosLGPD
        return Response(
            {
                "dimensao": list(
                    m.Dimensao.choices
                ),  # [["GPV","Gestão de privacidade"], ...]
                "criticidade": list(
                    m.Criticidade.choices
                ),  # [["NA","Não aplicável"], ...]
                "status": list(m.Status.choices),  # [["NA","Não aplicável"], ...]
            }
        )

    @action(detail=True, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def upload(self, request, pk=None):
        doc = self.get_object()
        # Permissão adicional já é validada pelo permission_classes em métodos de escrita,
        # mas mantemos uma checagem explícita aqui para upload:
        if request.user.role not in ["admin", "dpo"]:
            return Response(
                {"detail": "Sem permissão."},
                status=status.HTTP_403_FORBIDDEN,
            )

        file_obj = request.FILES.get("arquivo")
        if not file_obj:
            return Response(
                {"detail": "Envie o arquivo no campo 'arquivo'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        doc.arquivo = file_obj
        doc.save()
        return Response(self.get_serializer(doc).data, status=status.HTTP_200_OK)


# ViewSet para gerenciar o checklist da LGPD
class ChecklistViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Checklist.objects.all().order_by("id")
    serializer_class = ChecklistSerializer
    audit_module = "checklist"

    # === Filtro habilitado ===
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["is_completed"]

    # Esta função controla as permissões de acesso para cada ação da API
    def get_permissions(self):
        # Permite que Administradores e DPOs criem, alterem e excluam itens
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAdminOrDPO()]  # admin ou dpo
        return [permissions.IsAuthenticated()]  # leitura para autenticados


# ViewSet para InventarioDados
class InventarioDadosViewSet(AuditLogMixin, viewsets.ModelViewSet):
    lookup_value_regex = r"\d+"
    queryset = InventarioDados.objects.select_related(
        "criado_por"
    ).order_by(  # evita N+1 no autor
        "-data_criacao"
    )  # default
    serializer_class = InventarioDadosSerializer
    permission_classes = [SimpleRolePermission]
    pagination_class = DefaultPagination
    audit_module = "inventario"

    # Campo que identifica o "dono" do registro (para escopo 'own')
    OWN_FIELD = "criado_por"

    # Regras por ação e papel:
    # - Admin & DPO: CRUD total
    # - Gerente: listar/ver/criar; editar apenas os próprios; NUNCA excluir
    ROLE_PERMS = {
        "list": {"admin": "any", "dpo": "any", "gerente": "any"},
        "retrieve": {"admin": "any", "dpo": "any", "gerente": "any"},
        "create": {"admin": "any", "dpo": "any", "gerente": "any"},
        "update": {"admin": "any", "dpo": "any", "gerente": "own"},
        "partial_update": {"admin": "any", "dpo": "any", "gerente": "own"},
        "destroy": {"admin": "any", "dpo": "any", "gerente": None},
        "export_csv": {"admin": "any", "dpo": "any", "gerente": "any"},
        "export_xlsx": {"admin": "any", "dpo": "any", "gerente": "any"},
        "export_pdf": {"admin": "any", "dpo": "any", "gerente": "any"},
    }

    # Busca e ordenação no backend (DRF)
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    search_fields = [
        "unidade",
        "setor",
        "responsavel_email",
        "processo_negocio",
        "finalidade",
        "dados_pessoais",
    ]
    ordering_fields = [
        "id",
        "unidade",
        "setor",
        "responsavel_email",
        "processo_negocio",
        "data_criacao",
    ]
    ordering = ["-data_criacao"]  # default (reforço)

    filterset_fields = {
        "unidade": ["exact", "icontains"],
        "setor": ["exact", "icontains"],
        "responsavel_email": ["exact", "icontains"],
        "processo_negocio": ["exact", "icontains"],
        "controlador_operador": ["exact"],
        "impresso": ["exact"],
        "dados_menores": ["exact"],
        "transferencia_terceiros": ["exact"],
        "transferencia_internacional": ["exact"],
        "tipo_dado": ["exact"],
        "formato": ["exact"],
        "adequado_contratualmente": ["exact"],
    }

    SIMPLE_FILTERS = {
        "unidade": "unidade",
        "setor": "setor",
        "responsavel_email": "responsavel_email",
        "processo_negocio": "processo_negocio",
        "controlador_operador": "controlador_operador",
        "impresso": "impresso",
        "dados_menores": "dados_menores",
        "transferencia_terceiros": "transferencia_terceiros",
        "transferencia_internacional": "transferencia_internacional",
        "tipo_dado": "tipo_dado",
        "formato": "formato",
        "adequado_contratualmente": "adequado_contratualmente",
    }

    def get_queryset(self):
        qs = super().get_queryset()

        # aplica “escopo” por papel/ação (admin/dpo = any; gerente = own para update/partial_update)
        user = self.request.user
        role = getattr(user, "role", "") or ""
        action = getattr(self, "action", "list")
        scope = self.ROLE_PERMS.get(action, {}).get(role) or self.ROLE_PERMS.get(
            action, {}
        ).get("admin")

        if scope == "own":
            qs = qs.filter(**{self.OWN_FIELD: user})

        # filtros exatos pelos query params
        params = self.request.query_params
        for qp, field in self.SIMPLE_FILTERS.items():
            v = params.get(qp)
            if v is None or v == "":
                continue
            if field in {
                "impresso",
                "dados_menores",
                "transferencia_terceiros",
                "transferencia_internacional",
                "adequado_contratualmente",
            }:
                # normaliza 'sim'/'nao' -> True/False
                v = str(v).strip().lower()
                if v in ("true", "1", "sim", "yes"):
                    v = True
                elif v in ("false", "0", "nao", "não", "no"):
                    v = False
                else:
                    continue
            qs = qs.filter(**{field: v})

        return qs

    def perform_create(self, serializer):
        # Define automaticamente quem criou
        serializer.save(criado_por=self.request.user)

    def _export_cols_and_rows(self, qs):
        """Reaproveita cabeçalhos e formatação (bool/data) para todos os formatos."""
        cols = [
            ("id", "ID"),
            ("unidade", "Unidade"),
            ("setor", "Setor"),
            ("responsavel_email", "Responsável (E-mail)"),
            ("processo_negocio", "Processo de Negócio"),
            ("finalidade", "Finalidade"),
            ("dados_pessoais", "Dados Pessoais"),
            ("tipo_dado", "Tipo de Dado"),
            ("origem", "Origem"),
            ("formato", "Formato"),
            ("impresso", "Impresso"),
            ("titulares", "Titulares"),
            ("dados_menores", "Dados de menores"),
            ("base_legal", "Base Legal"),
            ("pessoas_acesso", "Pessoas com Acesso"),
            ("atualizacoes", "Atualizações (Quando)"),
            ("transmissao_interna", "Transmissão Interna"),
            ("transmissao_externa", "Transmissão Externa"),
            ("local_armazenamento_digital", "Local Armazenamento (Digital)"),
            ("controlador_operador", "Controlador/Operador"),
            ("motivo_retencao", "Motivo Retenção"),
            ("periodo_retencao", "Período Retenção"),
            ("exclusao", "Exclusão"),
            ("forma_exclusao", "Forma Exclusão"),
            ("transferencia_terceiros", "Transf. a Terceiros"),
            ("quais_dados_transferidos", "Quais Dados Transferidos"),
            ("transferencia_internacional", "Transf. Internacional"),
            ("empresa_terceira", "Empresa Terceira"),
            ("adequado_contratualmente", "Adequado Contratualmente"),
            ("paises_tratamento", "Países Tratamento"),
            ("medidas_seguranca", "Medidas de Segurança"),
            ("consentimentos", "Consentimentos"),
            ("observacao", "Observação"),
            ("criado_por", "Criado por (email)"),
            ("data_criacao", "Data Criação"),
            ("data_atualizacao", "Última Atualização"),
        ]

        br_tz = ZoneInfo("America/Sao_Paulo")
        rows = []
        for obj in qs.iterator():
            row = []
            for field, _ in cols:
                if field == "criado_por":
                    val = getattr(getattr(obj, "criado_por", None), "email", "") or ""
                else:
                    val = getattr(obj, field, "")

                if isinstance(val, bool):
                    val = "Sim" if val else "Não"
                elif isinstance(val, datetime.datetime):
                    if timezone.is_naive(val):
                        val = timezone.make_aware(val, timezone.get_default_timezone())
                    val = timezone.localtime(val, br_tz).strftime("%d/%m/%Y %H:%M")
                elif isinstance(val, datetime.date):
                    val = val.strftime("%d/%m/%Y")
                row.append(val if val is not None else "")
            rows.append(row)
        return cols, rows

    @staticmethod
    def _timestamp_br():
        br_tz = ZoneInfo("America/Sao_Paulo")
        dt = timezone.localtime(timezone.now(), timezone=br_tz)
        return f"{dt.day:02d}-{dt.month:02d}-{dt.year}_{dt.hour:02d}h{dt.minute:02d}min"

    @action(detail=False, methods=["get"], url_path=r"export/csv")
    def export_csv(self, request):
        qs = self.filter_queryset(self.get_queryset())
        cols, rows = self._export_cols_and_rows(qs)
        headers = [label for _, label in cols]

        file_name = f"inventarios-{self._timestamp_br()}.csv"
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        resp.write("\ufeff")  # BOM para Excel abrir acentos certinho

        writer = csv.writer(resp, lineterminator="\n")
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return resp

    @action(detail=False, methods=["get"], url_path=r"export/xlsx")
    def export_xlsx(self, request):
        qs = self.filter_queryset(self.get_queryset())
        cols, rows = self._export_cols_and_rows(qs)
        headers = [label for _, label in cols]

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventários"

        header_font = Font(bold=True)
        fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        for c, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        for col_idx, label in enumerate(headers, start=1):
            width = max(10, min(50, len(str(label)) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(output)
        output.seek(0)

        file_name = f"inventarios-{self._timestamp_br()}.xlsx"
        resp = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        return resp

    @action(detail=False, methods=["get"], url_path=r"export/pdf")
    def export_pdf(self, request):
        qs = self.filter_queryset(self.get_queryset())

        # Mapeamento de campos (reutiliza a mesma ordem dos outros exports)
        field_map = [
            ("id", "ID"),
            ("unidade", "Unidade"),
            ("setor", "Setor"),
            ("responsavel_email", "Responsável (E-mail)"),
            ("processo_negocio", "Processo de Negócio"),
            ("finalidade", "Finalidade"),
            ("dados_pessoais", "Dados Pessoais"),
            ("tipo_dado", "Tipo de Dado"),
            ("origem", "Origem"),
            ("formato", "Formato"),
            ("impresso", "Impresso"),
            ("titulares", "Titulares"),
            ("dados_menores", "Dados de menores"),
            ("base_legal", "Base Legal"),
            ("pessoas_acesso", "Pessoas com Acesso"),
            ("atualizacoes", "Atualizações (Quando)"),
            ("transmissao_interna", "Transmissão Interna"),
            ("transmissao_externa", "Transmissão Externa"),
            ("local_armazenamento_digital", "Local Armazenamento (Digital)"),
            ("controlador_operador", "Controlador/Operador"),
            ("motivo_retencao", "Motivo Retenção"),
            ("periodo_retencao", "Período Retenção"),
            ("exclusao", "Exclusão"),
            ("forma_exclusao", "Forma Exclusão"),
            ("transferencia_terceiros", "Transf. a Terceiros"),
            ("quais_dados_transferidos", "Quais Dados Transferidos"),
            ("transferencia_internacional", "Transf. Internacional"),
            ("empresa_terceira", "Empresa Terceira"),
            ("adequado_contratualmente", "Adequado Contratualmente"),
            ("paises_tratamento", "Países Tratamento"),
            ("medidas_seguranca", "Medidas de Segurança"),
            ("consentimentos", "Consentimentos"),
            ("observacao", "Observação"),
            ("criado_por", "Criado por (email)"),
            ("data_criacao", "Data Criação"),
            ("data_atualizacao", "Última Atualização"),
        ]

        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles["Heading3"]
        title_style.spaceAfter = 6

        label_style = ParagraphStyle(
            "Label",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            spaceAfter=0,
            spaceBefore=0,
        )
        value_style = ParagraphStyle(
            "Value",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            spaceAfter=0,
            spaceBefore=0,
        )

        # Helpers de formatação
        br_tz = ZoneInfo("America/Sao_Paulo")

        def fmt_val(v):
            if isinstance(v, bool):
                return "Sim" if v else "Não"
            if isinstance(v, datetime.datetime):
                if timezone.is_naive(v):
                    v = timezone.make_aware(v, timezone.get_default_timezone())
                return timezone.localtime(v, br_tz).strftime("%d/%m/%Y %H:%M")
            if isinstance(v, datetime.date):
                return v.strftime("%d/%m/%Y")
            return "" if v is None else str(v)

        # Montagem do PDF (cartões)
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=24,
            bottomMargin=24,
        )

        story = []
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors

        # Larguras: (Label, Value, Label, Value)
        col_widths = [90, 300, 110, 300]  # ~800pt úteis em A4 paisagem com margens

        for idx, obj in enumerate(qs.iterator(), start=1):
            # Título do cartão
            titulo = f"Inventário #{getattr(obj, 'id', '')}"
            subtitulo_bits = []
            if getattr(obj, "unidade", ""):
                subtitulo_bits.append(str(obj.unidade))
            if getattr(obj, "setor", ""):
                subtitulo_bits.append(str(obj.setor))
            if subtitulo_bits:
                titulo += f" — {' / '.join(subtitulo_bits)}"

            story.append(Paragraph(escape(titulo), title_style))
            story.append(Spacer(1, 4))

            # Constrói pares (label, value)
            pairs = []
            for field, label in field_map:
                if field == "criado_por":
                    v = getattr(getattr(obj, "criado_por", None), "email", "") or ""
                else:
                    v = getattr(obj, field, "")
                pairs.append(
                    (
                        Paragraph(escape(label), label_style),
                        Paragraph(
                            escape(fmt_val(v)).replace("\n", "<br/>"), value_style
                        ),
                    )
                )

            # Converte para linhas com 2 pares por linha -> 4 colunas
            data = []
            it = iter(pairs)
            for p in it:
                row = [p[0], p[1]]  # label, value
                try:
                    p2 = next(it)
                    row.extend([p2[0], p2[1]])
                except StopIteration:
                    # linha ímpar: completa com vazio
                    row.extend(["", ""])
                data.append(row)

            table = Table(data, colWidths=col_widths, repeatRows=0)
            table.setStyle(
                TableStyle(
                    [
                        # grade fina
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        # rótulos com leve fundo
                        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0F4FF")),
                        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F0F4FF")),
                        # alinhamentos & padding
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                        # quebra de linha até para palavras longas
                        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
                    ]
                )
            )

            story.append(table)

            # página nova por registro (mais legível)
            story.append(PageBreak())

        # Remove o último PageBreak se existir
        if story and isinstance(story[-1], PageBreak):
            story.pop()

        doc.build(story)

        # Nome do arquivo “03-09-2025_16h33min”
        dt_now = timezone.localtime(timezone.now(), timezone=br_tz)
        ts = f"{dt_now.day:02d}-{dt_now.month:02d}-{dt_now.year}_{dt_now.hour:02d}h{dt_now.minute:02d}min"
        file_name = f"inventarios-{ts}.pdf"

        pdf_bytes = buffer.getvalue()
        buffer.close()

        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        return resp


# ViewSet para MatrizRisco
class RiskViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Risk.objects.all().select_related("probabilidade", "impacto", "eficacia")
    serializer_class = RiskSerializer
    permission_classes = [IsAdminOrDPO]
    audit_module = "riscos"

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params

        setor = params.get("setor")
        if setor:
            qs = qs.filter(setor__icontains=setor)

        processo = params.get("processo")
        if processo:
            qs = qs.filter(processo__icontains=processo)

        matriz = params.get("matriz_filial")
        if matriz:
            qs = qs.filter(matriz_filial__icontains=matriz)

        return qs

    # filtros/busca/ordenação padrão
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "matriz_filial": ["exact", "icontains"],
        # "setor": ["exact", "icontains"],
        "processo": ["exact", "icontains"],
        "risco_residual": ["exact"],
        "tipo_controle": ["exact"],
        "probabilidade": ["exact"],
        "impacto": ["exact"],
    }
    search_fields = ["risco_fator", "processo", "setor", "matriz_filial"]
    ordering_fields = ["pontuacao", "criado_em", "atualizado_em"]
    ordering = ["-criado_em"]

    def perform_create(self, serializer):
        # não passar 'criado_por' (o modelo Risk não tem esse campo)
        serializer.save()

    @action(detail=False, methods=["get"], url_path=r"ranking/export/xlsx")
    def export_ranking_xlsx(self, request):
        qs = self.filter_queryset(self.get_queryset()).order_by(
            "-pontuacao", "-impacto__value", "-probabilidade__value", "-criado_em"
        )

        headers = [
            "ID",
            "Matriz/Filial",
            "Setor",
            "Processo",
            "Risco e Fator de Risco",
            "Prob (valor)",
            "Prob (rótulo)",
            "Impacto (valor)",
            "Impacto (rótulo)",
            "Pontuação",
            "Risco Residual",
            "Medidas de Controle",
            "Tipo Ctrl (C/D)",
            "Eficácia (rótulo)",
            "Resposta ao Risco",
            "Criado em",
        ]

        br_tz = ZoneInfo("America/Sao_Paulo")
        rows = []
        for r in qs.iterator():
            criado = (
                timezone.localtime(r.criado_em, br_tz).strftime("%d/%m/%Y %H:%M")
                if r.criado_em
                else ""
            )
            rows.append(
                [
                    r.id,
                    r.matriz_filial,
                    r.setor,
                    r.processo,
                    r.risco_fator,
                    getattr(getattr(r, "probabilidade", None), "value", None),
                    getattr(getattr(r, "probabilidade", None), "label_pt", None),
                    getattr(getattr(r, "impacto", None), "value", None),
                    getattr(getattr(r, "impacto", None), "label_pt", None),
                    r.pontuacao,
                    r.risco_residual,
                    r.medidas_controle or "",
                    r.tipo_controle or "",
                    getattr(getattr(r, "eficacia", None), "label_pt", None) or "",
                    r.resposta_risco or "",
                    criado,
                ]
            )

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking de Riscos"

        header_font = Font(bold=True)
        fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        for c, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # largura de colunas + filtro + congelar cabeçalho
        for col_idx, label in enumerate(headers, start=1):
            width = max(12, min(60, len(str(label)) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

        wb.save(output)
        output.seek(0)

        dt = timezone.localtime(timezone.now(), timezone=br_tz)
        file_name = f"ranking-riscos_{dt:%d-%m-%Y_%Hh%Mmin}.xlsx"

        resp = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        return resp

    @action(detail=False, methods=["get"], url_path=r"ranking/export/csv")
    def export_ranking_csv(self, request):
        qs = self.filter_queryset(self.get_queryset()).order_by(
            "-pontuacao", "-impacto__value", "-probabilidade__value", "-criado_em"
        )

        headers = [
            "ID",
            "Matriz/Filial",
            "Setor",
            "Processo",
            "Risco e Fator de Risco",
            "Prob (valor)",
            "Prob (rótulo)",
            "Impacto (valor)",
            "Impacto (rótulo)",
            "Pontuação",
            "Risco Residual",
            "Medidas de Controle",
            "Tipo Ctrl (C/D)",
            "Eficácia (rótulo)",
            "Resposta ao Risco",
            "Criado em",
        ]

        br_tz = ZoneInfo("America/Sao_Paulo")

        def rows_iter():
            yield headers
            for r in qs.iterator():
                criado = (
                    timezone.localtime(r.criado_em, br_tz).strftime("%d/%m/%Y %H:%M")
                    if r.criado_em
                    else ""
                )
                yield [
                    r.id,
                    r.matriz_filial,
                    r.setor,
                    r.processo,
                    r.risco_fator,
                    getattr(getattr(r, "probabilidade", None), "value", None),
                    getattr(getattr(r, "probabilidade", None), "label_pt", None),
                    getattr(getattr(r, "impacto", None), "value", None),
                    getattr(getattr(r, "impacto", None), "label_pt", None),
                    r.pontuacao,
                    r.risco_residual,
                    r.medidas_controle or "",
                    r.tipo_controle or "",
                    getattr(getattr(r, "eficacia", None), "label_pt", None) or "",
                    r.resposta_risco or "",
                    criado,
                ]

        # CSV em memória (com BOM p/ Excel abrir acentos)
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        br_now = timezone.localtime(timezone.now(), timezone=br_tz)
        file_name = f"ranking-riscos_{br_now:%d-%m-%Y_%Hh%Mmin}.csv"
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        resp.write("\ufeff")  # BOM

        writer = csv.writer(resp, lineterminator="\n")
        for row in rows_iter():
            writer.writerow(row)
        return resp

    @action(detail=False, methods=["get"], url_path=r"heatmap/export/csv")
    def export_heatmap_csv(self, request):
        qs = self.filter_queryset(self.get_queryset())
        buckets = {}
        total = 0
        for r in qs:
            if not (r.probabilidade_id and r.impacto_id):
                continue
            p = r.probabilidade.value
            i = r.impacto.value
            key = f"{p}-{i}"
            buckets[key] = buckets.get(key, 0) + 1
            total += 1

        headers = ["Probabilidade (1-5)", "Impacto (1-5)", "Contagem"]
        pairs = []
        # ordenar por prob desc e impacto desc (como heatmap “de cima à direita”)
        for p in range(5, 0, -1):
            for i in range(5, 0, -1):
                key = f"{p}-{i}"
                pairs.append((p, i, buckets.get(key, 0)))

        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        br_tz = ZoneInfo("America/Sao_Paulo")
        br_now = timezone.localtime(timezone.now(), timezone=br_tz)
        file_name = f"heatmap-riscos_{br_now:%d-%m-%Y_%Hh%Mmin}.csv"
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        resp.write("\ufeff")  # BOM

        writer = csv.writer(resp, lineterminator="\n")
        writer.writerow(headers)
        for p, i, c in pairs:
            writer.writerow([p, i, c])
        writer.writerow([])
        writer.writerow(["Total", "", total])
        return resp

    @action(detail=False, methods=["get"], url_path=r"ranking/export/pdf")
    def export_ranking_pdf(self, request):
        """
        Exporta o Ranking de Riscos em PDF (A4 paisagem) com tabela paginada.
        Respeita os mesmos filtros do list().
        """
        # --- consulta (mesma ordenação do ranking) ---
        qs = self.filter_queryset(self.get_queryset()).order_by(
            "-pontuacao", "-impacto__value", "-probabilidade__value", "-criado_em"
        )

        # --- cabeçalhos e linhas ---
        headers = [
            "ID",
            "Matriz/Filial",
            "Setor",
            "Processo",
            "Risco e Fator de Risco",
            "Prob",
            "Impacto",
            "Pontuação",
            "Risco Residual",
            "Medidas de Controle",
            "Tipo Ctrl",
            "Eficácia",
            "Resposta ao Risco",
            "Criado em",
        ]

        br_tz = ZoneInfo("America/Sao_Paulo")
        rows = []
        for r in qs.iterator():
            criado = (
                timezone.localtime(r.criado_em, br_tz).strftime("%d/%m/%Y %H:%M")
                if r.criado_em
                else ""
            )
            rows.append(
                [
                    str(r.id or ""),
                    r.matriz_filial or "",
                    r.setor or "",
                    r.processo or "",
                    r.risco_fator or "",
                    # Prob / Impact com rótulo (valor)
                    f"{getattr(getattr(r, 'probabilidade', None), 'label_pt', '')} ({getattr(getattr(r, 'probabilidade', None), 'value', '')})",
                    f"{getattr(getattr(r, 'impacto', None), 'label_pt', '')} ({getattr(getattr(r, 'impacto', None), 'value', '')})",
                    str(r.pontuacao or ""),
                    (r.risco_residual or "").capitalize(),
                    r.medidas_controle or "",
                    r.tipo_controle or "",
                    getattr(getattr(r, "eficacia", None), "label_pt", "") or "",
                    r.resposta_risco or "",
                    criado,
                ]
            )

        # --- montar PDF ---
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=22,
            bottomMargin=22,
        )

        styles = getSampleStyleSheet()
        title = Paragraph("Ranking de Riscos", styles["Heading2"])

        # Larguras aproximadas por coluna (em pontos; ~ 800 úteis em A4 paisagem com margens)
        # Ajuste fino se quiser mais/menos espaço em alguma coluna
        col_widths = [
            30,  # ID
            90,  # Matriz/Filial
            80,  # Setor
            100,  # Processo
            200,  # Risco e Fator
            80,  # Prob
            90,  # Impacto
            65,  # Pontuação
            80,  # Residual
            140,  # Medidas de Controle
            60,  # Tipo Ctrl
            90,  # Eficácia
            150,  # Resposta ao Risco
            90,  # Criado em
        ]

        # Converter linhas em células com quebra de linha
        body_style = styles["Normal"]
        body_style.fontSize = 8
        body_style.leading = 10

        header_style = ParagraphStyle(
            "TblHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            alignment=1,  # center
        )

        # Cabeçalho como Paragraph
        header_row = [Paragraph(h, header_style) for h in headers]

        table_data = [header_row]
        for row in rows:
            table_data.append(
                [Paragraph(escape(str(cell)), body_style) for cell in row]
            )

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6F0FF")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("LEADING", (0, 1), (-1, -1), 10),
                    # padding suave
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    # cabeçalho fixo (repeatRows=1 já cuida), mas realça borda
                    ("LINEBEFORE", (0, 0), (0, -1), 0.25, colors.grey),
                    ("LINEAFTER", (-1, 0), (-1, -1), 0.25, colors.grey),
                ]
            )
        )

        # Quebra automática em múltiplas páginas graças ao flowable Table
        story = [title, Spacer(1, 8), table]

        try:
            doc.build(story)
        except LayoutError:
            # fallback simples: remove título se estourar (raro), e tenta novamente
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=18,
                rightMargin=18,
                topMargin=12,
                bottomMargin=18,
            )
            story = [table]
            doc.build(story)

        pdf_bytes = buffer.getvalue()
        buffer.close()

        # Nome do arquivo
        dt_now = timezone.localtime(timezone.now(), timezone=br_tz)
        file_name = f"ranking-riscos_{dt_now:%d-%m-%Y_%Hh%Mmin}.pdf"

        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{file_name}; filename=\"{file_name}\""
        )
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        return resp

    @action(detail=False, methods=["get"], url_path=r"stats/by-band")
    def stats_by_band(self, request):
        """
        Contagem por faixa de risco residual (baixo/medio/alto/critico).
        """
        qs = self.filter_queryset(self.get_queryset())
        counter = Counter((r.risco_residual or "").lower() for r in qs)
        data = {
            "baixo": counter.get("baixo", 0),
            "medio": counter.get("medio", 0),
            "alto": counter.get("alto", 0),
            "critico": counter.get("critico", 0),
        }
        return Response(data)


class RankingRiscoViewSet(AuditLogMixin, viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]
    audit_module = "ranking_riscos"

    def list(self, request):
        self._log(request, "ACCESS")

        qs = Risk.objects.all().order_by(
            "-pontuacao", "-impacto__value", "-probabilidade__value", "-criado_em"
        )

        data = []
        for r in qs:
            data.append(
                {
                    "id": r.id,
                    "matriz_filial": r.matriz_filial,
                    "setor": r.setor,
                    "processo": r.processo,
                    "risco_fator": r.risco_fator,
                    "probabilidade": (
                        {
                            "value": getattr(r.probabilidade, "value", None),
                            "label": getattr(r.probabilidade, "label_pt", None),
                        }
                        if r.probabilidade_id
                        else None
                    ),
                    "impacto": (
                        {
                            "value": getattr(r.impacto, "value", None),
                            "label": getattr(r.impacto, "label_pt", None),
                        }
                        if r.impacto_id
                        else None
                    ),
                    "pontuacao": r.pontuacao,
                    "risco_residual": r.risco_residual,
                    "medidas_controle": r.medidas_controle or "",
                    "tipo_controle": r.tipo_controle or "",
                    "eficacia_label": getattr(
                        getattr(r, "eficacia", None), "label_pt", ""
                    )
                    or "",
                    "resposta_risco": r.resposta_risco or "",
                }
            )

        return Response(data)


# ViewSet para PlanoAcao
class ActionPlanViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    ViewSet de Planos de Ação, refletindo exatamente os campos do modelo.
    """

    queryset = ActionPlan.objects.all().select_related("risco")
    serializer_class = ActionPlanSerializer
    permission_classes = [IsAdminOrDPO]
    audit_module = "plano-acao"

    # ===== Filtros =====
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    filterset_fields = {
        "status": ["exact"],
        "prazo": ["exact", "lte", "gte"],
        "risco": ["exact"],  # FK direta
    }

    search_fields = [
        "como",
        "responsavel_execucao",
        "risco__risco_fator",  # busca textual dentro do risco relacionado
    ]

    ordering_fields = ["prazo", "id"]
    ordering = ["prazo"]

    def list(self, request):
        self._log(request, "ACCESS")

    # ===== Criação =====
    def perform_create(self, serializer):
        """
        Cria um novo plano de ação vinculado a um risco.
        """
        serializer.save()

    # ===== Atualização parcial (PATCH) =====
    def partial_update(self, request, *args, **kwargs):
        """
        Força PATCH a aceitar payloads parciais (sem todos os campos).
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    # ===== Limpar Complemento =====
    @action(detail=True, methods=["patch"], url_path="limpar")
    def limpar_complemento(self, request, pk=None):
        """
        Limpa os campos do plano sem excluir o registro.
        """
        plan = self.get_object()
        plan.como = ""
        plan.responsavel_execucao = ""
        plan.prazo = None
        plan.status = "nao_iniciado"
        plan.save(update_fields=["como", "responsavel_execucao", "prazo", "status"])
        return Response({"detail": "Complemento removido com sucesso."}, status=200)

    # ===== Estatísticas de atrasos =====
    @action(detail=False, methods=["get"], url_path="stats/overdue")
    def stats_overdue(self, request):
        """
        Planos de ação atrasados (não concluídos e com prazo vencido).
        """
        today = timezone.localdate()
        qs = self.filter_queryset(self.get_queryset()).filter(
            status__in=["nao_iniciado", "andamento"], prazo__lt=today
        )

        items = []
        for ap in qs.select_related("risco"):
            dias = (today - ap.prazo).days
            items.append(
                {
                    "id": ap.id,
                    "risco_id": ap.risco_id,
                    "risco_risco_fator": getattr(ap.risco, "risco_fator", ""),
                    "matriz_filial": getattr(ap.risco, "matriz_filial", ""),
                    "setor": getattr(ap.risco, "setor", ""),
                    "processo": getattr(ap.risco, "processo", ""),
                    "prazo": ap.prazo.strftime("%Y-%m-%d"),
                    "dias_atraso": dias,
                    "status": ap.status,
                }
            )

        return Response({"count": len(items), "items": items})

    @action(detail=False, methods=["get"], url_path="navegacao")
    def registrar_navegacao(self, request):
        page = request.GET.get("page", "")
        detalhe = f"Navegação página {page}" if page else "Navegação interna"

        self.extra_log_detail = detalhe
        return Response({"detail": "ok"})


class ActionPlanControleView(AuditLogMixin, viewsets.ReadOnlyModelViewSet):
    """
    Endpoint SOMENTE para a tela Controle de Ações.

    Devolve exatamente o mesmo payload de /riscos/ (RiskSerializer),
    só que com módulo de auditoria próprio: 'controle_plano_acao'.
    """

    queryset = (
        Risk.objects.select_related("probabilidade", "impacto", "eficacia")
        .prefetch_related("planos")  # garante os complementos
        .all()
    )
    serializer_class = RiskSerializer
    permission_classes = [IsAdminOrDPO]
    audit_module = "controle_plano_acao"
    pagination_class = None
    audit_ignore_models = ["Risk"]


class MonitoringActionViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = MonitoringAction.objects.all()
    serializer_class = MonitoringActionSerializer
    permission_classes = [IsAdminOrDPO]
    audit_module = "acao-monitoramento"

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    search_fields = ["framework_requisito", "responsavel", "escopo"]
    ordering_fields = ["data_monitoramento", "data_conclusao", "id"]
    ordering = ["-data_monitoramento"]


class IncidentViewSet(AuditLogMixin, viewsets.ModelViewSet):
    queryset = Incident.objects.all()
    serializer_class = IncidentSerializer
    permission_classes = [IsAdminOrDPO]
    audit_module = "incidentes"

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    search_fields = ["descricao", "fonte", "responsavel_analise", "decisoes_resolucao"]
    ordering_fields = ["data_registro", "data_encerramento", "numero_registro"]
    ordering = ["-data_registro"]


class RiskConfigView(APIView):
    """
    Retorna a parametrização usada pela Matriz de Riscos.
    GET /risk-config/
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import (
            LikelihoodItem,
            ImpactItem,
            ControlEffectivenessItem,
            RiskLevelBand,
            Instruction,
        )
        from .serializers import (
            LikelihoodItemSerializer,
            ImpactItemSerializer,
            ControlEffectivenessItemSerializer,
            RiskLevelBandSerializer,
            InstructionSerializer,
        )

        likelihood = LikelihoodItem.objects.all().order_by("value")
        impact = ImpactItem.objects.all().order_by("value")
        effectiveness = ControlEffectivenessItem.objects.all().order_by("value")
        bands = RiskLevelBand.objects.all().order_by("min_score")
        instructions = Instruction.objects.all().order_by("updated_at")

        data = {
            "likelihood": LikelihoodItemSerializer(likelihood, many=True).data,
            "impact": ImpactItemSerializer(impact, many=True).data,
            "effectiveness": ControlEffectivenessItemSerializer(
                effectiveness, many=True
            ).data,
            "bands": RiskLevelBandSerializer(bands, many=True).data,
            "instructions": InstructionSerializer(instructions, many=True).data,
        }
        return Response(data)


class HeatmapRiscoViewSet(AuditLogMixin, viewsets.ViewSet):
    permission_classes = [IsAdminOrDPO]

    audit_module = "heatmap_riscos"

    def list(self, request):
        self._log(request, "ACCESS")
        """
        Retorna:
        - riscos: lista completa (compatível com o Ranking)
        - buckets: contagem por prob-impact
        - grid, points
        - axes: labels PT de probabilidade/impacto
        """
        qs = Risk.objects.select_related("probabilidade", "impacto").all()

        # --- LISTA COMPLETA DE RISCOS (compatível com ranking-riscos) ---
        riscos_list = []
        for r in qs:
            riscos_list.append(
                {
                    "id": r.id,
                    "matriz_filial": r.matriz_filial,
                    "setor": r.setor,
                    "processo": r.processo,
                    "risco_fator": r.risco_fator,
                    "probabilidade": (
                        {
                            "value": getattr(r.probabilidade, "value", None),
                            "label": getattr(r.probabilidade, "label_pt", None),
                        }
                        if r.probabilidade_id
                        else None
                    ),
                    "impacto": (
                        {
                            "value": getattr(r.impacto, "value", None),
                            "label": getattr(r.impacto, "label_pt", None),
                        }
                        if r.impacto_id
                        else None
                    ),
                    "pontuacao": r.pontuacao,
                    "risco_residual": r.risco_residual,
                }
            )

        # --- buckets ---
        buckets = {}
        for r in qs:
            if not (r.probabilidade_id and r.impacto_id):
                continue
            p = r.probabilidade.value
            i = r.impacto.value
            key = f"{p}-{i}"
            buckets[key] = buckets.get(key, 0) + 1

        # --- grid / points ---
        size = 5
        grid = [[0 for _ in range(size + 1)] for _ in range(size + 1)]
        points = []

        for key, count in buckets.items():
            p_str, i_str = key.split("-")
            p, i = int(p_str), int(i_str)

            if 1 <= p <= 5 and 1 <= i <= 5:
                grid[p][i] = count
                points.append({"prob": p, "impact": i, "count": count})

        # --- eixos ---
        probs = list(
            LikelihoodItem.objects.order_by("value").values("value", "label_pt")
        )
        imps = list(ImpactItem.objects.order_by("value").values("value", "label_pt"))

        return Response(
            {
                "riscos": riscos_list,  # <- PARA O FRONTEND ATUAL
                "buckets": buckets,
                "grid": grid,
                "points": points,
                "axes": {"probabilidade": probs, "impacto": imps},
                "total": qs.count(),
            }
        )


# ===========================
# Esqueci minha senha / Reset
# ===========================

User = get_user_model()
_token_generator = PasswordResetTokenGenerator()


class PasswordResetRequestView(APIView):
    """
    POST { "email": "alguem@example.com" }
    Sempre responde 200 (não revela existência do e-mail).
    Se o e-mail existir, envia link com uid/token para o FRONTEND_URL.
    """

    permission_classes = [AllowAny]

    def post(self, request):
        email = (request.data.get("email") or "").strip().lower()

        # resposta neutra (evita enumeração de e-mails)
        ok = Response(
            {"detail": "Link enviado para o e-mail informado."},
            status=status.HTTP_200_OK,
        )
        if not email:
            return ok

        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            return ok  # não revela

        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = _token_generator.make_token(user)

        frontend = getattr(settings, "FRONTEND_URL", "http://127.0.0.1:5173").rstrip(
            "/"
        )
        reset_url = f"{frontend}/reset-password?uid={uid}&token={token}"

        send_html_email(
            subject="Redefinição de senha - Camaleão",
            to_email=user.email,
            template_name="emails/password_reset",  # backend/templates/emails/password_reset.(html|txt)
            context={"reset_url": reset_url},
        )
        return ok


class PasswordResetConfirmView(APIView):
    """
    POST { "uid": "...", "token": "...", "new_password": "...", "new_password2": "..." }
    Valida token e define a nova senha.
    """

    permission_classes = [AllowAny]

    def post(self, request):
        uid = request.data.get("uid") or ""
        token = request.data.get("token") or ""
        p1 = request.data.get("new_password") or ""
        p2 = request.data.get("new_password2") or ""

        if not (uid and token and p1 and p2):
            return Response(
                {"detail": "Dados incompletos."}, status=status.HTTP_400_BAD_REQUEST
            )
        if p1 != p2:
            return Response(
                {"new_password2": "As senhas não coincidem."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # resolve usuário
        try:
            user_id = force_str(urlsafe_base64_decode(uid))
            user = User.objects.get(pk=user_id)
        except Exception:
            return Response(
                {"detail": "Link inválido."}, status=status.HTTP_400_BAD_REQUEST
            )

        # valida token
        if not _token_generator.check_token(user, token):
            return Response(
                {"detail": "Link inválido ou expirado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # valida regras de senha do Django
        try:
            password_validation.validate_password(p1, user)
        except Exception as e:
            return Response(
                {"new_password": [str(err) for err in e.error_list]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # aplica nova senha
        user.set_password(p1)
        user.save()

        return Response(
            {"detail": "Senha redefinida com sucesso."}, status=status.HTTP_200_OK
        )


class InviteSetPasswordView(APIView):
    """
    Fluxo de convite (conta recém-criada sem senha):
      - GET  /auth/set-password/?uid=...&token=...
           -> valida o par uid/token (front decide se mostra o formulário)
      - POST /auth/set-password/
           body: { "uid": "...", "token": "...", "password": "...", "password2": "..." }
           -> define a senha se token válido.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        uidb64 = request.query_params.get("uid") or ""
        token = request.query_params.get("token") or ""
        if not (uidb64 and token):
            return Response(
                {"detail": "Parâmetros ausentes."}, status=status.HTTP_400_BAD_REQUEST
            )

        # Resolve usuário a partir do uid
        try:
            user_id = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=user_id)
        except Exception:
            return Response(
                {"detail": "Link inválido."}, status=status.HTTP_400_BAD_REQUEST
            )

        # Valida token (expiração considerada pelo generator)
        if not _token_generator.check_token(user, token):
            return Response(
                {"detail": "Token inválido ou expirado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({"detail": "Token válido."}, status=status.HTTP_200_OK)

    def post(self, request):
        uidb64 = request.data.get("uid") or ""
        token = request.data.get("token") or ""
        p1 = request.data.get("password") or ""
        p2 = request.data.get("password2") or ""

        if not (uidb64 and token and p1 and p2):
            return Response(
                {"detail": "Dados incompletos."}, status=status.HTTP_400_BAD_REQUEST
            )
        if p1 != p2:
            return Response(
                {"password2": "As senhas não coincidem."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Resolve usuário
        try:
            user_id = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=user_id)
        except Exception:
            return Response(
                {"detail": "Link inválido."}, status=status.HTTP_400_BAD_REQUEST
            )

        # Valida token
        if not _token_generator.check_token(user, token):
            return Response(
                {"detail": "Token inválido ou expirado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Valida força da senha pelas regras do Django (mantém consistência)
        try:
            password_validation.validate_password(p1, user)
        except Exception as e:
            # Converte mensagens dos validadores em uma saída amigável
            msgs = getattr(e, "messages", None)
            return Response(
                {"password": (msgs[0] if msgs else str(e))},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Define a senha e salva
        user.set_password(p1)
        # Se você quiser ativar a conta somente após definir a senha:
        # user.is_active = True
        user.save()

        return Response(
            {"detail": "Senha definida com sucesso."}, status=status.HTTP_200_OK
        )


class CalendarEventViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    Permite listar, criar, editar e excluir eventos do calendário.
    Apenas o usuário autenticado vê e gerencia seus próprios eventos.
    """

    serializer_class = CalendarEventSerializer
    permission_classes = [permissions.IsAuthenticated]
    audit_module = "calendario"

    def get_queryset(self):
        return CalendarEvent.objects.filter(user=self.request.user).order_by(
            "date", "time"
        )


class EnsureOverdueUpdatedView(APIView):
    permission_classes = [IsAdminOrDPO]  # ou [] se desejar público

    def post(self, request, *args, **kwargs):
        updated = update_overdue_actions_if_needed(force=False)
        return Response(
            {"updated": updated, "detail": "Overdue check executed (if needed)."}
        )


class LoginActivityViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Listagem de logins realizados no sistema.
    Acesso restrito a Admin e DPO.
    """

    queryset = LoginActivity.objects.all()
    serializer_class = LoginActivitySerializer
    permission_classes = [IsAdminOrDPO]

    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = ["email", "ip_address", "data_login"]
    ordering_fields = ["data_login", "email"]
    search_fields = ["email", "ip_address", "user_agent"]


class UserActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Listagem de ações executadas no sistema.
    Acesso restrito a Admin e DPO.
    """

    queryset = UserActivityLog.objects.all()
    serializer_class = UserActivityLogSerializer
    permission_classes = [IsAdminOrDPO]

    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        "email",
        "modulo",
        "operacao",
        "resultado",
        "ip",
        "timestamp",
    ]
    ordering_fields = [
        "timestamp",
        "email",
        "modulo",
        "operacao",
    ]
    search_fields = [
        "email",
        "modulo",
        "detalhe",
    ]
